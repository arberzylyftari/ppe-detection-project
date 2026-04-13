from flask import Flask, render_template, Response, jsonify, request, send_file
from ultralytics import YOLO
import cv2
import time
import os
import datetime
import threading
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

app = Flask(__name__)

MODEL_PATH = 'best.pt'
model = None

def load_model():
    global model
    print("Loading YOLOv8 model...")
    model = YOLO(MODEL_PATH)
    print("Model loaded!")

CLASSES = ['helmet', 'no-helmet', 'vest', 'no-vest',
           'gloves', 'no-gloves', 'boots', 'no-boots', 'person']

VIOLATION_CLASSES = {'no-helmet', 'no-vest', 'no-gloves', 'no-boots'}

CLASS_COLORS = {
    'helmet':     '#00ff88',
    'no-helmet':  '#ff4444',
    'vest':       '#00ff88',
    'no-vest':    '#ff4444',
    'gloves':     '#00ff88',
    'no-gloves':  '#ff4444',
    'boots':      '#00ff88',
    'no-boots':   '#ff4444',
    'person':     '#4488ff',
}

# ── Global State ───────────────────────────────────────────────────────────────
camera         = None
video_source   = None
video_name     = 'webcam'
use_webcam     = True
is_streaming   = True
video_ended    = False   # NEW: signals frontend video is done

current_stats = {
    'fps': 0,
    'inference_ms': 0,
    'detections': {c: 0 for c in CLASSES},
    'violation': False,
    'violation_count': 0,
    'total_frames': 0,
    'screenshots_saved': 0,
    'video_ended': False,
}

# Per-frame log for Excel report
frame_log      = []   # [{timestamp, frame, violations, detections...}]
violation_history  = []
detection_history  = []
stats_lock     = threading.Lock()
last_screenshot_time = 0

os.makedirs('screenshots', exist_ok=True)
os.makedirs('reports', exist_ok=True)
os.makedirs('uploads', exist_ok=True)


def get_camera():
    global camera
    if camera is None or not camera.isOpened():
        camera = cv2.VideoCapture(0)
        camera.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)
        camera.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
    return camera


def generate_frames():
    global current_stats, violation_history, detection_history
    global last_screenshot_time, video_ended

    if use_webcam:
        cap = get_camera()
    else:
        cap = cv2.VideoCapture(video_source)

    prev_time   = time.time()
    frame_count = 0
    video_ended = False
    with stats_lock:
        current_stats['video_ended'] = False

    while True:
        if not is_streaming:
            time.sleep(0.05)
            continue

        success, frame = cap.read()

        if not success:
            if not use_webcam:
                # Video finished - do NOT restart
                video_ended = True
                with stats_lock:
                    current_stats['video_ended'] = True
                break
            else:
                break

        # Skip every other frame for uploaded videos
        frame_count += 1
        if not use_webcam and frame_count % 2 != 0:
            continue

        # Resize for speed
        h_orig, w_orig = frame.shape[:2]
        scale = min(640 / w_orig, 480 / h_orig, 1.0)
        if scale < 1.0:
            frame = cv2.resize(frame, (int(w_orig * scale), int(h_orig * scale)))

        # YOLO inference
        start   = time.time()
        results = model(frame, conf=0.4, verbose=False)
        inference_ms = (time.time() - start) * 1000

        current_time = time.time()
        fps = 1.0 / max(current_time - prev_time, 0.001)
        prev_time = current_time

        # Parse detections
        detections = {c: 0 for c in CLASSES}
        violation  = False

        for result in results:
            for box in result.boxes:
                cls_id = int(box.cls[0])
                if cls_id < len(CLASSES):
                    cls_name = CLASSES[cls_id]
                    detections[cls_name] += 1
                    if cls_name in VIOLATION_CLASSES:
                        violation = True

        # Annotate frame
        annotated = results[0].plot()
        h, w = annotated.shape[:2]
        overlay  = annotated.copy()

        # Responsive banner - scales with frame width (handles portrait phone videos)
        banner_h   = max(36, int(h * 0.09))
        font_scale = max(0.45, min(1.1, w / 480))
        font_thick = max(1, int(font_scale * 1.8))

        if violation:
            label = 'VIOLATION DETECTED'
            color = (0, 0, 200)
        else:
            label = 'ALL CLEAR'
            color = (0, 160, 0)

        cv2.rectangle(overlay, (0, 0), (w, banner_h), color, -1)
        cv2.addWeighted(overlay, 0.65, annotated, 0.35, 0, annotated)
        (tw, th), _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_DUPLEX, font_scale, font_thick)
        tx = max(6, (w - tw) // 2)
        ty = int(banner_h * 0.72)
        cv2.putText(annotated, label, (tx, ty),
                    cv2.FONT_HERSHEY_DUPLEX, font_scale, (255, 255, 255), font_thick)

        # Auto screenshot
        if violation and (current_time - last_screenshot_time) > 5:
            ts   = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            path = f'screenshots/violation_{ts}.jpg'
            cv2.imwrite(path, annotated)
            last_screenshot_time = current_time
            with stats_lock:
                current_stats['screenshots_saved'] += 1

        # Update stats + frame log
        now_str = datetime.datetime.now().strftime('%H:%M:%S')
        with stats_lock:
            current_stats['fps']           = round(fps, 1)
            current_stats['inference_ms']  = round(inference_ms, 1)
            current_stats['detections']    = detections.copy()
            current_stats['violation']     = violation
            current_stats['total_frames'] += 1
            if violation:
                current_stats['violation_count'] += 1

            vcount = sum(detections[c] for c in VIOLATION_CLASSES)
            violation_history.append({'time': now_str, 'value': vcount})
            if len(violation_history) > 60:
                violation_history.pop(0)
            detection_history.append({'time': now_str, 'detections': detections.copy()})
            if len(detection_history) > 60:
                detection_history.pop(0)

            # Log every 5th frame for report (keeps file small)
            if current_stats['total_frames'] % 5 == 0:
                frame_log.append({
                    'timestamp':       now_str,
                    'frame':           current_stats['total_frames'],
                    'violation':       int(violation),
                    **{c: detections[c] for c in CLASSES},
                })

        ret, buffer = cv2.imencode('.jpg', annotated, [cv2.IMWRITE_JPEG_QUALITY, 80])
        yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')


# ── Excel Report Generator ─────────────────────────────────────────────────────
def generate_excel_report():
    wb = openpyxl.Workbook()

    # ── Colour helpers ────────────────────────────────────────────────────────
    def fill(hex_color):
        return PatternFill('solid', fgColor=hex_color.lstrip('#'))

    def bold(size=11, color='FFFFFF'):
        return Font(bold=True, size=size, color=color)

    thin = Side(style='thin', color='2A4A6F')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')

    DARK   = '0D1528'
    HEADER = '1E3A5F'
    GREEN  = '006644'
    RED    = 'AA1111'
    BLUE   = '0055AA'
    LIGHT  = '111C35'

    with stats_lock:
        snap_stats = current_stats.copy()
        snap_log   = list(frame_log)
        snap_vh    = list(violation_history)

    ts_now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    vname  = video_name if not use_webcam else 'Webcam Session'

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 1 – Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Summary'
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 18

    # Title row
    ws1.merge_cells('A1:C1')
    ws1['A1'] = '🦺  PPE Detection System – Analysis Report'
    ws1['A1'].font      = Font(bold=True, size=14, color='00AAFF')
    ws1['A1'].fill      = fill(DARK)
    ws1['A1'].alignment = center
    ws1.row_dimensions[1].height = 36

    # Meta rows
    meta = [
        ('Video / Source',  vname),
        ('Report Generated', ts_now),
        ('Total Frames Analysed', snap_stats['total_frames']),
        ('Violation Frames', snap_stats['violation_count']),
        ('Violation Rate',
         f"{round(snap_stats['violation_count']/max(snap_stats['total_frames'],1)*100,1)}%"),
        ('Screenshots Saved', snap_stats['screenshots_saved']),
        ('Avg FPS', snap_stats['fps']),
        ('Avg Inference', f"{snap_stats['inference_ms']} ms"),
    ]

    for i, (k, v) in enumerate(meta, start=2):
        ws1[f'A{i}'] = k
        ws1[f'B{i}'] = v
        ws1[f'A{i}'].font      = Font(bold=True, color='8AB4E0', size=10)
        ws1[f'B{i}'].font      = Font(color='E0E6F0', size=10)
        ws1[f'A{i}'].fill      = fill(LIGHT)
        ws1[f'B{i}'].fill      = fill(DARK)
        ws1[f'A{i}'].border    = border
        ws1[f'B{i}'].border    = border
        ws1.row_dimensions[i].height = 20

    # Class totals table header
    r = len(meta) + 3
    for col, label in enumerate(['Class', 'Total Detections', 'Type'], start=1):
        cell = ws1.cell(row=r, column=col, value=label)
        cell.font      = bold(10)
        cell.fill      = fill(HEADER)
        cell.alignment = center
        cell.border    = border
    ws1.row_dimensions[r].height = 22

    # Count totals from frame log
    class_totals = {c: 0 for c in CLASSES}
    for row in snap_log:
        for c in CLASSES:
            class_totals[c] += row.get(c, 0)

    for i, cls in enumerate(CLASSES):
        row_n = r + 1 + i
        is_viol = cls in VIOLATION_CLASSES
        ws1.cell(row=row_n, column=1, value=cls).font   = Font(color='E0E6F0', size=10)
        ws1.cell(row=row_n, column=2, value=class_totals[cls]).font = Font(color='00AAFF', bold=True, size=10)
        ws1.cell(row=row_n, column=3, value='⚠ VIOLATION' if is_viol else '✔ SAFE').font = Font(
            color='FF4444' if is_viol else '00FF88', bold=True, size=10)
        for col in range(1, 4):
            ws1.cell(row=row_n, column=col).fill   = fill(DARK if i % 2 == 0 else LIGHT)
            ws1.cell(row=row_n, column=col).border = border
            ws1.cell(row=row_n, column=col).alignment = center
        ws1.row_dimensions[row_n].height = 18

    # Bar chart – class totals
    chart = BarChart()
    chart.title        = 'Detections Per Class'
    chart.style        = 10
    chart.y_axis.title = 'Count'
    chart.x_axis.title = 'Class'
    chart.width        = 18
    chart.height       = 12

    data_ref   = Reference(ws1, min_col=2, min_row=r, max_row=r+len(CLASSES))
    labels_ref = Reference(ws1, min_col=1, min_row=r+1, max_row=r+len(CLASSES))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(labels_ref)
    ws1.add_chart(chart, f'E{r}')

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 2 – Frame Log (per-frame data for Power BI)
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Frame Log')
    ws2.sheet_view.showGridLines = False

    headers = ['Timestamp', 'Frame', 'Violation'] + CLASSES
    col_widths = [12, 8, 10] + [12] * len(CLASSES)

    for i, (h, w_) in enumerate(zip(headers, col_widths), start=1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w_
        cell = ws2.cell(row=1, column=i, value=h)
        cell.font      = bold(10)
        cell.fill      = fill(HEADER)
        cell.alignment = center
        cell.border    = border
    ws2.row_dimensions[1].height = 22

    for ri, row in enumerate(snap_log, start=2):
        vals = ([row['timestamp'], row['frame'], row['violation']] +
                [row.get(c, 0) for c in CLASSES])
        for ci, val in enumerate(vals, start=1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.fill      = fill(DARK if ri % 2 == 0 else LIGHT)
            cell.font      = Font(color='E0E6F0', size=9)
            cell.alignment = center
            cell.border    = border
        # Highlight violation rows red
        if row['violation']:
            for ci in range(1, len(headers)+1):
                ws2.cell(row=ri, column=ci).fill = fill('2A0A0A')

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 3 – Violation Timeline (for Power BI line chart)
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('Violation Timeline')
    ws3.sheet_view.showGridLines = False

    for ci, hdr in enumerate(['Time', 'Violation Count'], start=1):
        cell = ws3.cell(row=1, column=ci, value=hdr)
        cell.font = bold(10); cell.fill = fill(HEADER)
        cell.alignment = center; cell.border = border
    ws3.column_dimensions['A'].width = 14
    ws3.column_dimensions['B'].width = 16
    ws3.row_dimensions[1].height = 22

    for ri, entry in enumerate(snap_vh, start=2):
        ws3.cell(row=ri, column=1, value=entry['time']).font  = Font(color='8AB4E0', size=9)
        ws3.cell(row=ri, column=2, value=entry['value']).font = Font(color='FF4444', bold=True, size=9)
        for ci in range(1, 3):
            ws3.cell(row=ri, column=ci).fill   = fill(DARK if ri % 2 == 0 else LIGHT)
            ws3.cell(row=ri, column=ci).border = border
            ws3.cell(row=ri, column=ci).alignment = center

    # Save
    ts_file = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_name = vname.replace(' ', '_').replace('/', '_')
    filename = f'reports/PPE_Report_{safe_name}_{ts_file}.xlsx'
    wb.save(filename)
    return filename


# ── Routes ─────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html', classes=CLASSES, colors=CLASS_COLORS)


@app.route('/video_feed')
def video_feed():
    return Response(generate_frames(),
                    mimetype='multipart/x-mixed-replace; boundary=frame')


@app.route('/stats')
def stats():
    with stats_lock:
        return jsonify({
            'stats':             current_stats,
            'violation_history': list(violation_history[-30:]),
            'detection_history': list(detection_history[-1:]),
        })


@app.route('/upload_video', methods=['POST'])
def upload_video():
    global video_source, video_name, use_webcam, camera, frame_log, video_ended
    if 'video' not in request.files:
        return jsonify({'error': 'No file'}), 400
    f = request.files['video']
    path = os.path.join('uploads', f.filename)
    f.save(path)
    video_source = path
    video_name   = f.filename
    use_webcam   = False
    video_ended  = False
    frame_log    = []       # reset log for new video
    if camera:
        camera.release()
        camera = None
    with stats_lock:
        current_stats['video_ended'] = False
        current_stats['violation_count'] = 0
        current_stats['total_frames']    = 0
        current_stats['screenshots_saved'] = 0
    return jsonify({'success': True, 'path': path})


@app.route('/stop_stream', methods=['POST'])
def stop_stream():
    global is_streaming
    is_streaming = False
    return jsonify({'success': True})


@app.route('/start_stream', methods=['POST'])
def start_stream():
    global is_streaming
    is_streaming = True
    return jsonify({'success': True})


@app.route('/use_webcam', methods=['POST'])
def switch_to_webcam():
    global use_webcam, camera, frame_log, video_ended
    use_webcam  = True
    video_ended = False
    frame_log   = []
    if camera:
        camera.release()
        camera = None
    with stats_lock:
        current_stats['video_ended'] = False
    return jsonify({'success': True})


@app.route('/reset_stats', methods=['POST'])
def reset_stats():
    global violation_history, detection_history, frame_log, video_ended
    frame_log   = []
    video_ended = False
    with stats_lock:
        current_stats['violation_count']  = 0
        current_stats['total_frames']     = 0
        current_stats['screenshots_saved']= 0
        current_stats['video_ended']      = False
        violation_history.clear()
        detection_history.clear()
    return jsonify({'success': True})


@app.route('/generate_report', methods=['POST'])
def generate_report():
    try:
        filepath = generate_excel_report()
        return jsonify({'success': True, 'filename': os.path.basename(filepath)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/download_report/<filename>')
def download_report(filename):
    path = os.path.join('reports', filename)
    if os.path.exists(path):
        return send_file(path, as_attachment=True)
    return jsonify({'error': 'File not found'}), 404


@app.route('/screenshots')
def list_screenshots():
    files = sorted(os.listdir('screenshots'), reverse=True)[:10]
    return jsonify({'files': files})


if __name__ == '__main__':
    load_model()
    print("\n" + "="*50)
    print("  PPE Detection System running!")
    print("  Open: http://localhost:5000")
    print("="*50 + "\n")
    app.run(debug=False, host='0.0.0.0', port=5000, threaded=True)